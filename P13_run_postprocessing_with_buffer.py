# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from arcpy.sa import *
from datetime import datetime
import shutil
import os
import arcpy


def _eliminate_polygon_parts(path1, threshold_km2):
    # Reset ArcPy environment settings
    arcpy.ResetEnvironments()
    arcpy.env.workspace = None
    arcpy.env.outputCoordinateSystem = None
    arcpy.env.overwriteOutput = True

    # Get input shapefile folder and base name
    folder_path, file_name = os.path.split(path1)
    base_name = os.path.splitext(file_name)[0]

    # Define output paths
    projected_path = os.path.join(folder_path, base_name + "_v1_CEA.shp")
    eliminated_path = os.path.join(folder_path, base_name + "_v2_eli_CEA.shp")

    # Project to Cylindrical Equal Area (EPSG:54034)
    target_projection = arcpy.SpatialReference(54034)
    arcpy.Project_management(path1, projected_path, target_projection)

    # Convert threshold from square kilometers to square meters
    threshold_m2 = threshold_km2 * 1_000_000

    # Eliminate polygon parts based on area threshold
    arcpy.EliminatePolygonPart_management(
        in_features=projected_path,
        out_feature_class=eliminated_path,
        condition="AREA",
        part_area=threshold_m2,
        part_option="CONTAINED_ONLY"
    )

    # Reproject back to WGS 1984
    wgs84_projection = arcpy.SpatialReference(4326)
    final_path = os.path.join(folder_path, base_name + "_v3_eli_84.shp")
    arcpy.Project_management(eliminated_path, final_path, wgs84_projection)


def _extract_by_mask(tif_path, shp_path, output_result):
    # Enable overwrite
    arcpy.env.overwriteOutput = True

    # Extract raster by shapefile mask
    arcpy.sa.ExtractByMask(tif_path, shp_path).save(output_result)

    print(f"Extracted result saved to: {output_result}")


def _process_tif_with_fill_focal_analysis_5_5(input_tif, output_tif):
    """
    Process an input TIF using focal analysis and gap filling to generate a final raster.

    Steps
    -----
    1. Convert input raster to a reference raster (valid=1, NoData=0).
    2. Perform 5x5 focal sum analysis.
    3. Generate a processing raster based on neighborhood conditions.
    4. Fill NoData pixels using focal mean values.
    5. Remove all intermediate rasters.

    Parameters
    ----------
    input_tif : str
        Input TIF file path.
    output_tif : str
        Output TIF file path.
    """
    try:
        arcpy.env.overwriteOutput = True

        temp_dir = os.path.dirname(output_tif)
        fill_tif = os.path.join(temp_dir, "fill_tif.tif")
        reference_tif = os.path.join(temp_dir, "reference_tif.tif")
        focal_tif = os.path.join(temp_dir, "focal_tif.tif")

        reference_raster = arcpy.sa.Con(arcpy.sa.IsNull(input_tif), 0, 1)
        reference_raster.save(reference_tif)

        neighborhood = arcpy.sa.NbrRectangle(5, 5, "CELL")
        focal_sum = arcpy.sa.FocalStatistics(reference_tif, neighborhood, "SUM", "DATA")

        processed_raster = arcpy.sa.Con(focal_sum >= 4, 1, 0)
        processed_raster.save(focal_tif)

        tif_2 = FocalStatistics(input_tif, neighborhood, "MEAN", "DATA")
        tif_2.save(fill_tif)

        raster1 = Raster(input_tif)
        raster3 = Raster(focal_tif)
        raster2 = Raster(fill_tif)

        raster4 = Con(IsNull(raster1) & (raster3 == 1), raster2, raster1)
        raster4.save(output_tif)

        arcpy.management.Delete(reference_tif)
        arcpy.management.Delete(focal_tif)
        arcpy.management.Delete(fill_tif)

        print(f"Processing completed. Output saved to: {output_tif}")

    except Exception as e:
        print(f"An error occurred: {e}")


def _process_tif_with_focal_analysis_5_5(input_tif, output_tif):
    """
    Process an input TIF using 5x5 focal analysis to remove isolated pixels.

    Parameters
    ----------
    input_tif : str
        Input TIF file path.
    output_tif : str
        Output TIF file path.
    """
    try:
        arcpy.env.overwriteOutput = True

        temp_dir = os.path.dirname(output_tif)
        reference_tif = os.path.join(temp_dir, "reference_tif.tif")
        focal_tif = os.path.join(temp_dir, "focal_tif.tif")

        reference_raster = arcpy.sa.Con(arcpy.sa.IsNull(input_tif), 0, 1)
        reference_raster.save(reference_tif)

        neighborhood = arcpy.sa.NbrRectangle(5, 5, "CELL")
        focal_sum = arcpy.sa.FocalStatistics(reference_tif, neighborhood, "SUM", "DATA")

        processed_raster = arcpy.sa.Con(focal_sum >= 5, 1, 0)
        processed_raster.save(focal_tif)

        input_raster = Raster(input_tif)
        expression = "SetNull((\"{}\" == {}), \"{}\")".format(
            processed_raster, 0, input_raster
        )
        arcpy.gp.RasterCalculator_sa(expression, output_tif)

        arcpy.management.Delete(reference_tif)
        arcpy.management.Delete(focal_tif)

        print(f"Processing completed. Output saved to: {output_tif}")

    except Exception as e:
        print(f"An error occurred: {e}")


def _process_tif_with_focal_analysis_3_3(input_tif, output_tif):
    """
    Process an input TIF using 3x3 focal analysis to remove isolated pixels.

    Parameters
    ----------
    input_tif : str
        Input TIF file path.
    output_tif : str
        Output TIF file path.
    """
    try:
        arcpy.env.overwriteOutput = True

        temp_dir = os.path.dirname(output_tif)
        reference_tif = os.path.join(temp_dir, "reference_tif.tif")
        focal_tif = os.path.join(temp_dir, "focal_tif.tif")

        reference_raster = arcpy.sa.Con(arcpy.sa.IsNull(input_tif), 0, 1)
        reference_raster.save(reference_tif)

        neighborhood = arcpy.sa.NbrRectangle(3, 3, "CELL")
        focal_sum = arcpy.sa.FocalStatistics(reference_tif, neighborhood, "SUM", "DATA")

        processed_raster = arcpy.sa.Con(focal_sum >= 3, 1, 0)
        processed_raster.save(focal_tif)

        input_raster = Raster(input_tif)
        expression = "SetNull((\"{}\" == {}), \"{}\")".format(
            processed_raster, 0, input_raster
        )
        arcpy.gp.RasterCalculator_sa(expression, output_tif)

        arcpy.management.Delete(reference_tif)
        arcpy.management.Delete(focal_tif)

        print(f"Processing completed. Output saved to: {output_tif}")

    except Exception as e:
        print(f"An error occurred: {e}")


def _merge_and_dissolve_can_shapefiles(path4):
    path1 = os.path.join(path4, "01_Monthly_Drawdown_Area")
    path2 = os.path.join(path4, "09k_Can_Create_One_Area")

    if os.path.exists(path2):
        shutil.rmtree(path2)
    os.makedirs(path2)

    output_filename = "Can_Create_One_Area.shp"

    arcpy.env.workspace = path1
    shapefiles = arcpy.ListFeatureClasses("*.shp")

    if not shapefiles:
        print("No shapefiles found in the specified directory.")
        return

    merged_shp = os.path.join(path2, "Merged_Temp.shp")
    arcpy.management.Merge(shapefiles, merged_shp)

    dissolved_shp = os.path.join(path2, output_filename)
    arcpy.management.Dissolve(merged_shp, dissolved_shp, multi_part="MULTI_PART")

    arcpy.management.Delete(merged_shp)

    print(f"Merged and dissolved shapefile saved to: {dissolved_shp}")


def _calculate_area(file_path, shp_path):
    """
    Calculate area for a raster or shapefile.

    Returns
    -------
    float
        Area in square kilometers, or None if calculation fails.
    """
    arcpy.env.workspace = None
    arcpy.env.extent = None
    arcpy.env.outputCoordinateSystem = None
    arcpy.env.snapRaster = None

    file_extension = os.path.splitext(file_path)[1].lower()
    file_basename = os.path.splitext(os.path.basename(file_path))[0]

    if file_extension == '.shp':
        target_projection = arcpy.SpatialReference(54034)
        desc = arcpy.Describe(file_path)
        current_projection = desc.spatialReference

        if current_projection.name != target_projection.name:
            proj_shp_path = os.path.join(os.path.dirname(file_path), f"{file_basename}_projected.shp")
            arcpy.Project_management(file_path, proj_shp_path, target_projection)

            total_area = 0
            with arcpy.da.SearchCursor(proj_shp_path, ["SHAPE@AREA"]) as cursor:
                for row in cursor:
                    total_area += row[0] / 1_000_000

            arcpy.Delete_management(proj_shp_path)
            return total_area
        else:
            total_area = 0
            with arcpy.da.SearchCursor(file_path, ["SHAPE@AREA"]) as cursor:
                for row in cursor:
                    total_area += row[0] / 1_000_000
            return total_area

    elif file_extension == '.tif':
        raster = arcpy.Raster(file_path)
        spatial_ref = raster.spatialReference
        if spatial_ref is None or spatial_ref.type == 'Unknown':
            print(f"{file_path} has no defined spatial reference. Area calculation skipped.")
            return None

        cea_sr = arcpy.SpatialReference(54034)
        projected_tif = os.path.join(os.path.dirname(file_path), f"{file_basename}_projected.tif")
        if arcpy.Exists(projected_tif):
            arcpy.Delete_management(projected_tif)

        arcpy.ProjectRaster_management(file_path, projected_tif, cea_sr)

        binary_tif = os.path.join(os.path.dirname(file_path), f"{file_basename}_binary.tif")
        if arcpy.Exists(binary_tif):
            arcpy.Delete_management(binary_tif)

        projected_tif_raster = Raster(projected_tif)
        binary_tif_raster = SetNull(IsNull(projected_tif_raster), 1)
        binary_tif_raster.save(binary_tif)

        vector_output = os.path.join(os.path.dirname(file_path), f"{file_basename}_polygon.shp")
        if arcpy.Exists(vector_output):
            arcpy.Delete_management(vector_output)

        try:
            arcpy.conversion.RasterToPolygon(
                binary_tif,
                vector_output,
                "NO_SIMPLIFY",
                create_multipart_features="MULTIPLE_OUTER_PART"
            )
        except arcpy.ExecuteError:
            print(f"Raster-to-polygon conversion failed: {arcpy.GetMessages()}")
            return None

        clipped_output = os.path.join(os.path.dirname(file_path), f"{file_basename}_clipped.shp")
        if arcpy.Exists(clipped_output):
            arcpy.Delete_management(clipped_output)

        try:
            arcpy.analysis.Clip(vector_output, shp_path, clipped_output)
        except arcpy.ExecuteError:
            print(f"Vector clipping failed: {arcpy.GetMessages()}")
            return None

        total_area = 0
        with arcpy.da.SearchCursor(clipped_output, ["SHAPE@AREA"]) as cursor:
            for row in cursor:
                total_area += row[0] / 1_000_000

        arcpy.Delete_management(projected_tif)
        arcpy.Delete_management(binary_tif)
        arcpy.Delete_management(vector_output)
        arcpy.Delete_management(clipped_output)

        return total_area

    else:
        print("Unsupported file type.")
        return None


def P13_run_postprocessing_with_buffer(shp_path, path4, lakename, WAF_or_Not):
    """
    Run post-processing workflow including buffering, masking,
    focal analysis, raster calculations, and final DEM generation.
    """
    xlsx_path = os.path.join(path4, "Results.xlsx")

    tif_1_path0 = os.path.join(path4, '09_Initial_reconstructed_tif')
    tif_1_path = os.path.join(tif_1_path0, 'New_One_Tif.tif')

    base_path = os.path.join(path4, '09_Process_shps')
    tif_4_path = os.path.join(base_path, 'tif5.tif')
    final_output_path = os.path.join(path4, f'{lakename}_Final_DEM.tif')

    _merge_and_dissolve_can_shapefiles(path4)

    arcpy.CheckOutExtension("Spatial")
    arcpy.management.Delete(base_path)
    if not os.path.exists(base_path):
        os.makedirs(base_path)

    Can_Create_shp_path = os.path.join(
        path4, "09k_Can_Create_One_Area", "Can_Create_One_Area.shp"
    )

    final_v1_output_tif = os.path.join(path4, f'{lakename}_v1.tif')
    _process_tif_with_focal_analysis_3_3(tif_1_path, final_v1_output_tif)

    final_v2_output_tif = os.path.join(path4, f'{lakename}_v2.tif')
    _process_tif_with_focal_analysis_5_5(final_v1_output_tif, final_v2_output_tif)

    _process_tif_with_fill_focal_analysis_5_5(final_v2_output_tif, tif_4_path)

    _eliminate_polygon_parts(Can_Create_shp_path, 0.1)

    Can_Create_shp_path_eli = os.path.join(
        path4, "09k_Can_Create_One_Area", "Can_Create_One_Area_v3_eli_84.shp"
    )
    Can_Create_shp_path_proj = os.path.join(
        path4, "09k_Can_Create_One_Area", "Can_Create_One_Area_v2_eli_CEA.shp"
    )

    _extract_by_mask(tif_4_path, Can_Create_shp_path_eli, final_output_path)

    arcpy.management.Delete(base_path)
    shutil.rmtree(base_path, ignore_errors=True)

    Don_Scan_tif_Area = _calculate_area(tif_1_path, shp_path)
    Max_shp_Area = _calculate_area(shp_path, shp_path)
    Max_tif_Area = _calculate_area(final_output_path, shp_path)
    Can_Create_shp = _calculate_area(Can_Create_shp_path_proj, shp_path)

    percent1 = (Don_Scan_tif_Area / Can_Create_shp) * 100
    percent2 = (Max_tif_Area / Can_Create_shp) * 100
    percent3 = (Max_tif_Area / Max_shp_Area) * 100

    print(
        f"Initial ratio: {percent1:.2f}%, "
        f"Final ratio: {percent2:.2f}%, "
        f"Overall ratio: {percent3:.2f}% — Processing completed."
    )

    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Lake",
            "pre_scan/Can",
            "aft_scan/Can",
            "aft_scan/All",
            "Timestamp",
            "Area",
            "WAF_or_Not"
        ])
    else:
        wb = load_workbook(xlsx_path)
        ws = wb.active

    ws.append([
        lakename,
        percent1 / 100,
        percent2 / 100,
        percent3 / 100,
        current_datetime,
        Max_shp_Area,
        WAF_or_Not
    ])

    last_row = ws.max_row
    ws[f"B{last_row}"].number_format = '0.00%'
    ws[f"C{last_row}"].number_format = '0.00%'
    ws[f"D{last_row}"].number_format = '0.00%'

    wb.save(xlsx_path)

    arcpy.CheckInExtension("Spatial")
